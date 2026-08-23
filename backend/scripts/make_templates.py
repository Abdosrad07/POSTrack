"""Génère les 10 gabarits Excel officiels dans scripts/templates/.

Les colonnes sont issues de REQUIRED_COLUMNS du service de validation,
augmentées des colonnes optionnelles documentées — identiques à celles
servies par GET /api/partners/{id}/imports/templates/{entity_type}.

Usage (depuis backend/):
    .\\venv\\Scripts\\python scripts\\make_templates.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font

from app.services.import_validation_service import REQUIRED_COLUMNS

OUT_DIR = Path(__file__).resolve().parent / "templates"

OPTIONAL = {
    "DSM": ["zone"],
    "POS": ["address", "zone", "stock_initial"],
    "BTS": ["operateur", "technologie", "capacite_max", "latitude", "longitude"],
    "BTS_RELEVE": ["debit", "connexions", "latence", "date_releve"],
    "SIM": ["status"],
    "PRIME_PERIOD": ["statut"],
    "REQUETE": ["description", "priorite", "entite_en_charge", "nombre_demande"],
}

SAMPLES = {
    "PARTNER": {"code_partenaire": "PART-001", "name": "Camtel Express"},
    "DSM": {"matricule": "DSM-DLA-01", "full_name": "Jean Marc", "zone": "Douala Akwa"},
    "POS": {"code_pos": "POS-0001", "name": "Kiosque Akwa", "dsm_matricule": "DSM-DLA-01",
             "date_creation": "2026-01-15", "date_expiration": "2026-12-31",
             "address": "Boulevard de la Liberté", "zone": "Akwa", "stock_initial": 50},
    "BTS": {"code_bts": "BTS-DLA-01", "operateur": "CAMTEL", "technologie": "4G",
             "capacite_max": 1000, "latitude": 4.0511, "longitude": 9.7679},
    "BTS_RELEVE": {"bts_code": "BTS-DLA-01", "charge": 62.5, "taux_saturation": 48.0,
                    "rendement": 91.2, "debit": 12.4, "connexions": 320, "latence": 45,
                    "date_releve": "2026-08-01"},
    "SIM": {"iccid": "89237010000000000001", "pos_code": "POS-0001", "status": "EN_STOCK"},
    "PRIME_PERIOD": {"code": "P2026-T3", "label": "Trimestre 3", "start_date": "2026-07-01",
                      "end_date": "2026-09-30", "statut": "OPEN"},
    "PRIME": {"pos_code": "POS-0001", "prime_period_code": "P2026-T3", "montant": 15000},
    "REQUETE": {"external_id": "EXT-REQ-001", "type_requete": "AJOUT", "titre": "Ajout POS Akwa",
                 "description": "Besoin de 2 POS supplémentaires", "priorite": "NORMALE",
                 "entite_en_charge": "AC Bépanda", "nombre_demande": 2},
}


def make(entity_key: str) -> Path:
    required = REQUIRED_COLUMNS[entity_key]
    optional = [c for c in OPTIONAL.get(entity_key, []) if c not in required]
    columns = required + optional
    sample = SAMPLES.get(entity_key, {})

    wb = Workbook()
    ws = wb.active
    ws.title = entity_key
    ws.append(columns)
    ws.append([sample.get(c, "") for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, col in enumerate(columns, start=1):
        width = max(len(col), len(str(sample.get(col, "")))) + 4
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(width, 42)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / f"gabarit_{entity_key.lower()}.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    for key in REQUIRED_COLUMNS:
        print("Gabarit écrit :", make(key))
